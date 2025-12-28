import streamlit as st
from openpyxl import load_workbook
from PIL import Image, ImageDraw
from utils import create_final_bundle, generate_certificate_image, has_descenders, get_font

favicon = Image.open("favicon.png")

st.set_page_config(page_title="Certificate Generator", page_icon=favicon, layout="wide")

st.title("Certificate Generator")
st.markdown("Upload assets, calibrate positions, and download your certificates.")

st.divider()
st.subheader("1. Upload Assets")

col1, col2, col3 = st.columns(3)
with col1:
    template_file = st.file_uploader("Template (PNG/JPG)", type=["png", "jpg", "jpeg"])
with col2:
    font_file = st.file_uploader("Font (TTF/OTF)", type=["ttf", "otf"])
with col3:
    excel_file = st.file_uploader("Excel Data", type=["xlsx"])

st.divider()
st.subheader("2. Configuration")
col1, col2 = st.columns(2)
with col1:
    font_size = st.number_input("Font Size", value=110, step=5)
with col2:
    text_color = st.color_picker("Text Color", value='#000000')


if template_file and font_file and excel_file:

    template_img = Image.open(template_file)
    font_bytes = font_file.getvalue()
    loaded_font = get_font(font_bytes, font_size)
    
    col1, col2 = st.columns(2)
    with col1:
        captial_exceptions = st.text_input('Captial Exceptions', value='')
        st.write("Check which characters in your specific font dip below the line:")
    with col2:
        inspect_chars = "G J P Y Q F"
        
        char_map_width = 800
        char_map_height = 120
        char_map = Image.new("RGB", (char_map_width, char_map_height), "#FFFFFF")
        char_draw = ImageDraw.Draw(char_map)
        
        baseline_y = 80
        char_draw.line([(20, baseline_y), (char_map_width - 20, baseline_y)], fill="red", width=2)

        _, _, tw, th = char_draw.textbbox((0, 0), inspect_chars, font=loaded_font)
        
        char_draw.text(((char_map_width - tw) / 2, baseline_y - 100), inspect_chars, font=loaded_font, fill=text_color)

        st.image(char_map, width='stretch')
        
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.worksheets[0]
    
    names = [str(row[0]).strip() for row in ws.iter_rows(values_only=True) if row and row[0]]
    
    normal_list, desc_list = [], []
    for n in names:
        word_list = normal_list
        for word in n.lower().split(' '):
            if(word[0] in captial_exceptions or has_descenders(word[1:])):
                word_list = desc_list
                break
        word_list.append(n.title())
    
    st.divider()
    st.subheader("3. Live Preview")
    c1, c2 = st.columns(2)
    
    with c1:
        st.info(f"Normal Names: {len(normal_list)}")
        if normal_list:
            off_n = st.number_input("Offset (Normal)", value=0, step=5)
            img_norm = generate_certificate_image(normal_list[0], off_n, template_img, loaded_font, text_color)
            st.image(img_norm, caption=f"Preview: {normal_list[0]}", width='stretch')

    with c2:
        st.info(f"Descender Names: {len(desc_list)}")
        if desc_list:
            off_d = st.number_input("Offset (Descender)", value=0, step=5)
            img_desc = generate_certificate_image(desc_list[0], off_d, template_img, loaded_font, text_color)
            st.image(img_desc, caption=f"Preview: {desc_list[0]}", width='stretch')

    st.divider()
    st.subheader("4. Generate")
    if st.button("Generate ZIP", width='stretch'):
        with st.spinner("Building all certificates..."):
            final_zip = create_final_bundle(
                normal_list, desc_list, off_n, off_d, 
                template_img, loaded_font, text_color
            )
            
            st.success("All certificates generated successfully!")
            st.download_button(
                label="Download All Certificates (One ZIP)",
                data=final_zip,
                file_name="All_Certificates.zip",
                mime="application/zip"
            )
else:
    st.warning("* Please upload all three files to proceed.")